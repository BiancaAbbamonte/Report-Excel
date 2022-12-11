import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.chart import BarChart, Reference
import string


excel_file = pd.read_excel('supermarket_sales.xlsx')
report_table = excel_file.pivot_table(index='Gender', columns='Product line', values='Total', aggfunc='sum').round(0)
report_table.to_excel('report_2021.xlsx', sheet_name='Report', startrow=4)

wb = load_workbook('report_2021.xlsx')
sheet = wb['Report']
ws = wb.active
ws.sheet_view.showGridLines = False

min_column = wb.active.min_column
max_column = wb.active.max_column
min_row = wb.active.min_row
max_row = wb.active.max_row

alphabet = list(string.ascii_uppercase)
excel_alphabet = alphabet[0:max_column]

sheet['B7'] = '=SUM(B5:B6)'
sheet['B7'].style = 'Currency'

for i in excel_alphabet:
    if i!='A':
        sheet[f'{i}{max_row+1}'] = f'=SUM({i}{min_row+1}:{i}{max_row})'
        sheet[f'{i}{max_row+1}'].style = 'Currency'
# adding total label
sheet[f'{excel_alphabet[0]}{max_row+1}'] = 'Total'

# adding a chart
barchart = BarChart()
data = Reference(sheet, min_col=min_column + 1, max_col=max_column, min_row=min_row, max_row=max_row)  # including headers
categories = Reference(sheet, min_col=min_column, max_col=min_column, min_row=min_row + 1, max_row=max_row)  # not including headers
barchart.add_data(data, titles_from_data=True)
barchart.set_categories(categories)
sheet.add_chart(barchart, "B12")  # location chart
barchart.title = 'Sales by Product line'
barchart.style = 2

sheet['A1'] = 'Sales Report'
sheet['A2'] = '2021'
sheet['A1'].font = Font('Arial', bold=True, size=20)
sheet['A2'].font = Font('Arial', bold=True, size=10)

wb.save('report_2021.xlsx')
